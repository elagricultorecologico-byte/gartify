from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Casos de Uso"

COLOR_HEADER = "1E40AF"
COLOR_ALT    = "EFF6FF"
COLOR_BLANCO = "FFFFFF"

headers = [
    "No", "Nombre del caso de uso", "Actor principal", "Actores secundarios",
    "Descripcion", "Precondiciones", "Postcondiciones",
    "Flujo principal", "Flujo alternativo / Excepciones",
    "Entidades relacionadas", "Prioridad", "Modulo"
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

casos = [
    ("UC-001","Registrarse como cliente","Visitante","Sistema",
     "El visitante crea una cuenta de cliente con nombre, email y contrasena.",
     "El usuario no tiene cuenta registrada.",
     "Cuenta creada con rol CUSTOMER; sesion iniciada automaticamente.",
     "1. Accede a /registro\n2. Selecciona pestana Cliente\n3. Rellena nombre, email y contrasena\n4. Envia el formulario\n5. Sistema crea el usuario y redirige a /cuenta",
     "Email ya registrado -> error de validacion.\nContrasena debil -> error de validacion.",
     "User","Alta","Autenticacion"),

    ("UC-002","Registrarse como taller","Propietario de taller","Sistema",
     "El propietario crea una cuenta vinculando sus datos personales y los del taller.",
     "El usuario no tiene cuenta registrada.",
     "Cuenta con rol GARAGE_OWNER creada; registro Garage creado; redirige a /cuenta/taller.",
     "1. Accede a /registro\n2. Selecciona pestana Taller\n3. Rellena datos personales y del taller\n4. Envia el formulario\n5. Sistema crea User + Garage y redirige",
     "Email ya registrado -> error.\nCampos obligatorios vacios -> validacion Zod.",
     "User, Garage","Alta","Autenticacion"),

    ("UC-003","Iniciar sesion","Usuario registrado","Sistema, Auth.js",
     "El usuario accede a su area privada mediante email y contrasena.",
     "El usuario tiene una cuenta activa.",
     "Sesion JWT creada; redirigido segun rol.",
     "1. Accede a /login\n2. Introduce email y contrasena\n3. Sistema valida credenciales con bcryptjs\n4. Redirige segun rol: /cuenta, /cuenta/taller, /admin o /distribuidor/dashboard",
     "Credenciales incorrectas -> mensaje de error.\nCuenta inexistente -> mensaje de error.",
     "User","Alta","Autenticacion"),

    ("UC-004","Cerrar sesion","Usuario autenticado","Sistema",
     "El usuario finaliza su sesion activa.",
     "El usuario tiene sesion activa.",
     "Sesion destruida; redirige a /.",
     "1. Hace clic en Cerrar sesion\n2. Sistema invalida el token JWT\n3. Redirige a la home",
     "--",
     "User","Alta","Autenticacion"),

    ("UC-005","Buscar talleres","Visitante / Cliente","Sistema",
     "El usuario busca talleres filtrando por tipo de servicio y ciudad.",
     "Al menos un taller activo en la BD.",
     "Lista de talleres filtrada visible en lista y mapa.",
     "1. Accede a / o /talleres\n2. Selecciona servicio y ciudad en el buscador\n3. Sistema consulta talleres activos\n4. Muestra resultados en lista y mapa Leaflet",
     "Sin resultados -> mensaje vacio.\nCiudad no reconocida -> lista vacia.",
     "Garage, GarageService","Alta","Busqueda"),

    ("UC-006","Filtrar talleres por criterios","Visitante / Cliente","Sistema",
     "El usuario aplica filtros adicionales (verificado, servicios, coche de cortesia, recogida) sobre los resultados.",
     "Resultados de busqueda visibles.",
     "Lista actualizada segun filtros activos.",
     "1. Visualiza lista de talleres\n2. Activa filtros disponibles\n3. Sistema filtra en tiempo real",
     "Combinacion de filtros sin resultados -> mensaje vacio.",
     "Garage, GarageService","Alta","Busqueda"),

    ("UC-007","Ver perfil de taller","Visitante / Cliente","Sistema",
     "El usuario consulta informacion detallada de un taller: descripcion, servicios, horarios, valoraciones y ubicacion.",
     "El taller existe y esta activo.",
     "Informacion del taller mostrada correctamente.",
     "1. Hace clic en un taller de la lista\n2. Sistema carga /talleres/[id]\n3. Muestra servicios, horarios, mapa, resenas y boton de reserva",
     "Taller inactivo -> no aparece en busqueda.",
     "Garage, GarageService, GarageSchedule, Review","Alta","Busqueda"),

    ("UC-008","Ver mapa de talleres","Visitante / Cliente","Sistema, Leaflet",
     "El usuario visualiza los talleres en un mapa interactivo con marcadores naranjas.",
     "Al menos un taller con coordenadas lat/lng.",
     "Mapa Leaflet renderizado con marcadores.",
     "1. Accede a /talleres\n2. Sistema carga GarageMap (dynamic import, SSR disabled)\n3. Muestra marcadores por cada taller con lat/lng",
     "Taller sin coordenadas -> no aparece en mapa pero si en lista.",
     "Garage","Media","Busqueda"),

    ("UC-009","Crear reserva","Cliente autenticado","Sistema, Garage, GarageService",
     "El cliente reserva un servicio en un taller mediante un wizard de 3 pasos.",
     "Cliente autenticado; taller activo con servicios configurados.",
     "Reserva creada con estado PENDING; redirige a /reserva/[id].",
     "1. Accede a /talleres/[id]/reservar\n2. Paso 1: selecciona servicio\n3. Paso 2: selecciona fecha y franja horaria\n4. Paso 3: introduce datos del vehiculo\n5. Confirma -> POST /api/bookings\n6. Redirige a confirmacion",
     "Cliente no autenticado -> redirige a /login.\nSlot ya ocupado -> error.",
     "User, Garage, GarageService, Booking","Alta","Reservas"),

    ("UC-010","Ver mis reservas (cliente)","Cliente autenticado","Sistema",
     "El cliente consulta el listado de sus reservas con estado, taller, fecha y servicio.",
     "Cliente autenticado.",
     "Listado de reservas del cliente visible.",
     "1. Accede a /cuenta\n2. Sistema carga reservas del usuario via GET /api/bookings\n3. Muestra lista con estado y acciones",
     "Sin reservas -> mensaje informativo.",
     "User, Booking, Garage, GarageService","Alta","Reservas"),

    ("UC-011","Cancelar reserva","Cliente autenticado","Sistema",
     "El cliente cancela una reserva pendiente o confirmada.",
     "Reserva en estado PENDING o CONFIRMED.",
     "Estado de la reserva actualizado a CANCELLED.",
     "1. Accede a /cuenta\n2. Hace clic en Cancelar\n3. Sistema actualiza estado via PATCH /api/bookings/[id]",
     "Reserva ya completada o cancelada -> accion no disponible.",
     "Booking","Alta","Reservas"),

    ("UC-012","Ver confirmacion de reserva","Cliente autenticado","Sistema",
     "El cliente visualiza el resumen de su reserva tras completar el wizard.",
     "Reserva creada correctamente.",
     "Detalles de la reserva mostrados en pantalla.",
     "1. Sistema redirige a /reserva/[id]\n2. Muestra resumen: taller, servicio, fecha, precio y datos del vehiculo",
     "ID de reserva invalido -> error 404.",
     "Booking, Garage, GarageService","Alta","Reservas"),

    ("UC-013","Ver dashboard del taller","Propietario de taller","Sistema",
     "El propietario visualiza estadisticas y reservas recientes de su taller.",
     "Propietario autenticado con taller vinculado.",
     "Dashboard con stats y reservas recientes visible.",
     "1. Accede a /cuenta/taller\n2. Sistema carga stats: reservas hoy, pendientes, completadas\n3. Muestra listado de reservas",
     "Sin taller vinculado -> redirige a configuracion.",
     "Garage, Booking, GarageService","Alta","Portal Taller"),

    ("UC-014","Gestionar estado de reserva","Propietario de taller","Sistema",
     "El propietario cambia el estado de una reserva recibida.",
     "Propietario autenticado; reservas existentes en su taller.",
     "Estado de la reserva actualizado.",
     "1. Accede a /cuenta/taller\n2. Localiza la reserva\n3. Cambia estado via selector\n4. Sistema actualiza via PATCH /api/bookings/[id]",
     "Transicion de estado no permitida -> error.\nReserva de otro taller -> error 403.",
     "Booking, Garage","Alta","Portal Taller"),

    ("UC-015","Anadir servicio al taller","Propietario de taller","Sistema",
     "El propietario anade un nuevo servicio al catalogo de su taller indicando tipo, precio, duracion y tipos de vehiculo.",
     "Propietario autenticado con taller vinculado.",
     "Nuevo GarageService creado y visible en el catalogo.",
     "1. Accede a /cuenta/taller/servicios\n2. Rellena formulario: tipo, nombre, precio, duracion, tipos de vehiculo\n3. Envia -> POST /api/garage/services\n4. Lista se actualiza",
     "Ningun tipo de vehiculo seleccionado -> error de validacion.\nCampos obligatorios vacios -> validacion Zod.",
     "GarageService, Garage","Alta","Portal Taller"),

    ("UC-016","Eliminar servicio del taller","Propietario de taller","Sistema",
     "El propietario elimina un servicio de su catalogo.",
     "Servicio existente vinculado al taller del propietario.",
     "GarageService eliminado de la BD.",
     "1. Accede a /cuenta/taller/servicios\n2. Hace clic en eliminar\n3. Sistema elimina el registro",
     "Servicio con reservas futuras -> advertencia al propietario.",
     "GarageService","Media","Portal Taller"),

    ("UC-017","Editar perfil del taller","Propietario de taller","Sistema",
     "El propietario actualiza los datos publicos de su taller: nombre, descripcion, direccion, telefono, logo, servicios adicionales.",
     "Propietario autenticado con taller vinculado.",
     "Datos del taller actualizados en BD.",
     "1. Accede a /cuenta/taller/perfil\n2. Edita los campos deseados\n3. Guarda -> PATCH /api/garage/profile\n4. Sistema actualiza el registro Garage",
     "Direccion sin geocodificacion -> lat/lng no actualizados.",
     "Garage","Alta","Portal Taller"),

    ("UC-018","Gestionar ofertas del taller","Propietario de taller","Sistema",
     "El propietario crea y gestiona ofertas de horario con precio especial para atraer clientes.",
     "Propietario autenticado con taller vinculado.",
     "GarageOffer creada o eliminada.",
     "1. Accede a /cuenta/taller/ofertas\n2. Crea oferta indicando label, dias, horario y precio\n3. Sistema guarda la oferta activa",
     "Solapamiento de horarios -> permitido sin validacion adicional.",
     "GarageOffer, Garage","Media","Portal Taller"),

    ("UC-019","Contratar plan de suscripcion","Propietario de taller","Sistema, Stripe",
     "El propietario selecciona y contrata un plan PRO o PREMIUM a traves de Stripe Checkout.",
     "Propietario autenticado; plan actual STARTER.",
     "Sesion de Stripe Checkout iniciada; webhook actualiza Garage.plan tras el pago.",
     "1. Accede a /precios o /cuenta/taller/planes\n2. Selecciona plan y modalidad (mensual/anual)\n3. Sistema crea sesion Stripe Checkout\n4. Redirige a Stripe\n5. Tras pago exitoso, webhook actualiza plan y fechas",
     "Pago fallido -> Stripe gestiona el reintento.\nYa tiene suscripcion activa -> redirige al portal.",
     "Garage, User (stripeCustomerId, stripePriceId)","Alta","Suscripciones"),

    ("UC-020","Gestionar suscripcion en Stripe","Propietario de taller","Stripe",
     "El propietario accede al portal de Stripe para ver facturas, cambiar metodo de pago o cancelar su suscripcion.",
     "Propietario con suscripcion activa y stripeCustomerId.",
     "Portal de Stripe Customer cargado.",
     "1. Hace clic en Gestionar suscripcion\n2. Sistema crea sesion de Customer Portal\n3. Redirige al portal de Stripe",
     "Sin suscripcion activa -> boton no disponible.",
     "User (stripeCustomerId), Garage","Alta","Suscripciones"),

    ("UC-021","Ver panel de administracion","Administrador","Sistema",
     "El administrador accede al panel central con estadisticas globales y accesos a la gestion de la plataforma.",
     "Usuario autenticado con rol ADMIN.",
     "Panel de administracion visible.",
     "1. Accede a /admin\n2. Sistema verifica rol ADMIN\n3. Muestra stats globales y accesos a modulos",
     "Rol distinto a ADMIN -> redirige a /.",
     "User, Garage, Distributor","Alta","Administracion"),

    ("UC-022","Gestionar usuarios (admin)","Administrador","Sistema",
     "El administrador consulta, busca y modifica el rol de los usuarios registrados en la plataforma.",
     "Administrador autenticado.",
     "Listado de usuarios visible; rol actualizado si procede.",
     "1. Accede a /admin/usuarios\n2. Busca por email o nombre\n3. Cambia rol del usuario si es necesario\n4. Sistema actualiza el registro User",
     "Cambio de rol sobre la propia cuenta -> advertencia.",
     "User","Alta","Administracion"),

    ("UC-023","Gestionar talleres (admin)","Administrador","Sistema",
     "El administrador consulta, verifica y activa/desactiva talleres registrados.",
     "Administrador autenticado.",
     "Estado del taller actualizado (isActive, isVerified).",
     "1. Accede a /admin/talleres\n2. Ve listado con plan, estado y verificacion\n3. Activa, desactiva o verifica el taller",
     "--",
     "Garage","Alta","Administracion"),

    ("UC-024","Gestionar distribuidores (admin)","Administrador","Sistema",
     "El administrador crea distribuidores y los vincula a usuarios con rol DISTRIBUTOR.",
     "Administrador autenticado.",
     "Distributor creado y/o vinculado a User.",
     "1. Accede a /admin/distribuidores\n2. Crea distribuidor o selecciona existente\n3. Vincula a un usuario con rol DISTRIBUTOR via PATCH /api/admin/distributor/[id]/link-user\n4. Sistema actualiza Distributor.userId",
     "Usuario ya vinculado a otro distribuidor -> error.\nEmail duplicado -> error de unicidad.",
     "Distributor, User","Alta","Administracion"),

    ("UC-025","Ver dashboard del distribuidor","Distribuidor autenticado","Sistema",
     "El distribuidor visualiza sus KPIs (pedidos, facturacion, ticket medio) y la tabla de pedidos recibidos de los talleres.",
     "Usuario autenticado con rol DISTRIBUTOR; registro Distributor vinculado.",
     "Dashboard con stats, KPIs y tabla de pedidos cargado.",
     "1. Accede a /distribuidor/dashboard\n2. Sistema verifica rol y busca Distributor por userId\n3. Carga stats: total, pendientes, confirmados, en camino, entregados\n4. Carga KPIs: facturacion mes, ticket medio, pedidos vencidos\n5. Renderiza tabla de pedidos",
     "Sin registro Distributor vinculado -> mensaje de cuenta en configuracion.",
     "Distributor, PartOrder, PartOrderLine, Garage","Alta","Distribuidor"),

    ("UC-026","Ver y filtrar pedidos","Distribuidor autenticado","Sistema",
     "El distribuidor consulta los pedidos recibidos y aplica filtros de busqueda, estado, periodo y vencidos.",
     "Distribuidor autenticado con pedidos existentes.",
     "Listado de pedidos filtrado y paginado visible.",
     "1. Visualiza tabla en dashboard\n2. Busca por taller o numero de pedido\n3. Filtra por estado (pills)\n4. Filtra por periodo (hoy/semana/mes)\n5. Activa filtro rapido de vencidos (+24h)\n6. Navega por paginas",
     "Sin pedidos que coincidan -> mensaje diferenciado.",
     "PartOrder, PartOrderLine, Garage, DistributorPart, Part","Alta","Distribuidor"),

    ("UC-027","Cambiar estado de pedido","Distribuidor autenticado","Sistema",
     "El distribuidor avanza el estado de un pedido siguiendo el flujo: PENDING -> CONFIRMED -> SHIPPED -> DELIVERED (o CANCELLED).",
     "Pedido en estado activo (no DELIVERED ni CANCELLED).",
     "Estado actualizado; statusUpdatedAt renovado.",
     "1. Localiza el pedido en la tabla\n2. Selecciona nuevo estado en el selector Mover a...\n3. Sistema envia PATCH /api/distribuidor/orders/[id]\n4. Valida transicion permitida\n5. Actualiza status y statusUpdatedAt en BD",
     "Transicion no permitida -> error 400.\nPedido de otro distribuidor -> error 403.",
     "PartOrder, Distributor","Alta","Distribuidor"),

    ("UC-028","Identificar pedidos vencidos +24h","Distribuidor autenticado","Sistema",
     "El sistema identifica y resalta pedidos activos que llevan mas de 24h en su estado actual sin avanzar.",
     "Pedidos activos con statusUpdatedAt superior a 24h.",
     "Pedidos vencidos resaltados en ambar; alerta en dashboard; filtro rapido activo.",
     "1. Al cargar el dashboard, sistema cuenta pedidos vencidos\n2. Si > 0, muestra banner de alerta\n3. En tabla, filas con fondo ambar e icono de advertencia\n4. Boton filtro activa vista solo vencidos",
     "--",
     "PartOrder","Alta","Distribuidor"),

    ("UC-029","Ver detalle de lineas de pedido","Distribuidor autenticado","Sistema",
     "El distribuidor expande un pedido para ver las referencias, cantidades y precios de cada linea.",
     "Pedido con lineas existentes.",
     "Tabla de lineas del pedido visible.",
     "1. En la tabla hace clic en Detalle\n2. Se expande seccion con tabla de lineas\n3. Muestra referencia proveedor, descripcion, cantidad, precio unitario y subtotal",
     "Pedido sin lineas -> tabla vacia.",
     "PartOrderLine, DistributorPart, Part","Media","Distribuidor"),
]

thin   = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for row_idx, caso in enumerate(casos, 2):
    fill_color = COLOR_ALT if row_idx % 2 == 0 else COLOR_BLANCO
    for col_idx, value in enumerate(caso, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.fill      = PatternFill("solid", fgColor=fill_color)
        cell.border    = border
        cell.font      = Font(name="Calibri", size=9)

widths = [8, 28, 18, 20, 45, 35, 35, 55, 45, 30, 10, 16]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 30
for i in range(2, len(casos) + 2):
    ws.row_dimensions[i].height = 95

ws.freeze_panes = "A2"

path = r"d:\DescargasD\works\gartify\casos_de_uso_gartify.xlsx"
wb.save(path)
print(f"Guardado en: {path}")
