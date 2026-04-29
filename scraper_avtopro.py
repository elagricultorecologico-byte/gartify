"""
Scraper avtopro.es con Playwright — Marcas francesas
Salida: avtopro_francesas.xlsx
"""

import sys, time, re, json
sys.stdout.reconfigure(encoding="utf-8")

from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_URL = "https://avtopro.es"
MARCAS   = ["nissan"]  # Cambiar según la marca a scrapear
DELAY    = 1.0  # segundos entre navegaciones


# ─── Extracción ───────────────────────────────────────────────────────────────

def get_modelos(page, marca: str) -> list[dict]:
    base = f"{BASE_URL}/catalog/{marca}"
    page.goto(base, wait_until="networkidle", timeout=30000)
    page.wait_for_timeout(2000)

    links = page.query_selector_all("a")
    modelos = []
    seen = set()
    for a in links:
        href = a.get_attribute("href") or ""
        texto = a.inner_text().strip()
        # Hrefs relativos tipo ./clio/ o ./19/ (sin /model-)
        if (
            href.startswith("./")
            and "/model-" not in href
            and href not in seen
            and texto
        ):
            # Construir URL absoluta desde la base del catálogo
            slug = href.lstrip("./").rstrip("/")
            full = f"{base}/{slug}"
            seen.add(href)
            # Limpiar prefijo "Recambios/Repuestos para [Marca] "
            texto = re.sub(r'^(?:Recambios|Repuestos)\s+para\s+\S+\s+', '', texto, flags=re.IGNORECASE).strip()
            modelos.append({"nombre": texto, "url": full})
    return modelos


def get_variantes(page, modelo_url: str) -> list[dict]:
    page.goto(modelo_url, wait_until="networkidle", timeout=30000)
    page.wait_for_timeout(1500)

    links = page.query_selector_all("a")
    variantes = []
    seen = set()
    for a in links:
        href = a.get_attribute("href") or ""
        texto = a.inner_text().strip()
        if "model-" in href and href not in seen and texto:
            # Resolver URL absoluta
            if href.startswith("http"):
                full = href
            elif href.startswith("/"):
                full = BASE_URL + href
            else:
                # Relativa como ./model-27963/ o model-27963/
                full = modelo_url.rstrip("/") + "/" + href.lstrip("./")
            seen.add(href)
            # Limpiar prefijo "Repuestos [Marca] " o "Recambios [Marca] "
            texto = re.sub(r'^(?:Recambios|Repuestos)\s+\S+\s+', '', texto, flags=re.IGNORECASE).strip()

            # Intentar extraer años del contenedor padre (ej: "2013-2021" junto al nombre)
            anios_variante = ""
            try:
                padre = a.evaluate("el => el.closest('li, div, article, td, tr') ? el.closest('li, div, article, td, tr').innerText : ''")
                m = re.search(r'\b((?:19|20)\d{2})\s*[-–]\s*((?:19|20)\d{2})\b', padre or "")
                if m and int(m.group(1)) >= 1950:
                    anios_variante = f"{m.group(1)}-{m.group(2)}"
            except Exception:
                pass

            variantes.append({"nombre": texto, "url": full, "anios_variante": anios_variante})

    if not variantes:
        variantes.append({"nombre": "—", "url": modelo_url, "anios_variante": ""})
    return variantes


def get_anios_variante(page) -> str:
    """Extrae el año de producción de la variante desde la página ya cargada."""
    try:
        el = page.query_selector("[class*='year']")
        if el:
            txt = el.inner_text()
            m = re.search(r'((?:19|20)\d{2})\s*[-–]\s*((?:19|20)\d{2})', txt)
            if m and int(m.group(1)) >= 1950:
                return f"{m.group(1)}-{m.group(2)}"
        # Fallback: h2 con formato "(YYYY-YYYY)"
        h2 = page.query_selector("h2")
        if h2:
            txt = h2.inner_text()
            m = re.search(r'\((\d{4})-(\d{4})\)', txt)
            if m and int(m.group(1)) >= 1950:
                return f"{m.group(1)}-{m.group(2)}"
    except Exception:
        pass
    return ""


def parse_motores(page, variante_url: str) -> list[dict]:
    page.goto(variante_url, wait_until="networkidle", timeout=30000)
    page.wait_for_timeout(1500)

    # Intentar extraer filas de la tabla de motores
    motores = []

    # Buscar por selectores comunes
    rows = page.query_selector_all(
        ".catalog-model__engines-item, "
        ".engines-list__item, "
        "tr.engine-row, "
        "[class*='engine-item'], "
        "[class*='engines-item']"
    )

    # Fallback: todas las filas de tabla
    if not rows:
        rows = page.query_selector_all("table tr")

    for row in rows:
        # Preferir el atributo title del span interno (contiene datos completos)
        span = row.query_selector("span[title]")
        texto = (span.get_attribute("title") if span else None) or row.inner_text()
        texto = texto.strip()
        if not texto or len(texto) < 5:
            continue

        m = extraer_datos_motor(texto)
        if m:
            motores.append(m)

    # Segundo fallback: leer todo el texto de la sección de motores
    if not motores:
        seccion = page.query_selector("[class*='engines'], [class*='motor'], .catalog-model")
        if seccion:
            texto_completo = seccion.inner_text()
            for linea in texto_completo.splitlines():
                linea = linea.strip()
                if len(linea) < 8:
                    continue
                m = extraer_datos_motor(linea)
                if m:
                    motores.append(m)

    return motores


def extraer_datos_motor(texto: str) -> dict | None:
    """
    Formato típico: '1.5 dCi 90 K9K 608'  o  '0.9 TCE 75 (BHNP) H4B 408'
    Estructura:     [cilindrada] [nombre] [cv] [(variante)] [codigo]
    """
    import datetime
    anio_actual = datetime.date.today().year

    # Nombre del motor: parte antes del primer paréntesis o antes de los años
    nombre_limpio = re.split(r'\s*\((?:19|20)\d{2}', texto)[0].strip()
    # Quitar también sufijos de specs al final (Diésel, Gasolina, HP, kW, etc.)
    nombre_limpio = re.sub(r'\s+(Diésel|Gasolina|Diesel|Gasolina|GLP|Híbrido)\s*.*$', '', nombre_limpio, flags=re.IGNORECASE).strip()

    m = {
        "motor": nombre_limpio[:60],
        "codigo": "",
        "cilindrada": "",
        "cv": "",
        "kw": "",
        "combustible": "",
        "anios": "",
    }

    # Código motor: extraer desde el texto ANTES del paréntesis de años
    # Ej: "1.4 GA14DE (1995 - 2001) Gasolina..." → buscar en "1.4 GA14DE"
    pre_year = re.split(r'\s*\((?:19|20)\d{2}', texto)[0].strip()
    # Código de dos partes: "F8M 700", "C1E 715", "K9K 608"
    two_part = re.search(r'\b([A-Z][A-Z0-9]{1,5})\s+(\d{3,4})\s*$', pre_year)
    if two_part:
        m["codigo"] = two_part.group(1) + " " + two_part.group(2)
    else:
        # Código de una parte: "GA14DE", "CD20", "QG16DE", "YD22DDTi", "DV6FC"
        # Permite minúsculas al final (ej: YD22DDTi) pero exige al menos un dígito
        single = re.search(r'\b([A-Z][A-Za-z0-9]{2,8})\s*$', pre_year)
        if single and re.search(r'\d', single.group(1)):
            m["codigo"] = single.group(1)

    # Cilindrada: primer decimal del texto (ej: 1.5, 0.9, 2.0)
    cil = re.search(r'^(\d+[\.,]\d)', texto)
    if cil:
        m["cilindrada"] = cil.group(1).replace(",", ".")

    # CV: con unidad explícita o como número suelto antes del código
    cv_unidad = re.search(r'(\d{2,4})\s*(?:CV|HP|cv|hp|KM)\b', texto)
    if cv_unidad:
        m["cv"] = cv_unidad.group(1)
    else:
        # Número de 2-3 dígitos entre el nombre y el código: ej "dCi 90 K9K"
        cv_implicito = re.search(
            r'(?:TCe?|dCi|HDi|CDTi|TDi|VTi|THP|PureTech|BlueHDi|e-Tech|hybrid|LPG|SCe|GTe|GTi|D4D|D-4D|TFSI|TSI|TDCi|dTi|\d{1,2}[Vv])\s+(\d{2,4})\b',
            texto, re.IGNORECASE
        )
        if cv_implicito:
            m["cv"] = cv_implicito.group(1)
        else:
            # Último recurso: número suelto de 2-3 dígitos antes del código de motor
            cv_fallback = re.search(r'\b(\d{2,3})\s+(?:[A-Z][A-Z0-9]{2,6})', texto)
            if cv_fallback:
                m["cv"] = cv_fallback.group(1)

    # KW
    kw = re.search(r'(\d{2,4})\s*(?:KW|kW|kw)\b', texto)
    if kw:
        m["kw"] = kw.group(1)

    # Años desde el title: (1983 - 1989) — año fin = año actual significa "en producción"
    if not m["anios"]:
        anios2 = re.search(r'\((\d{4})\s*[-–]\s*(\d{4}|)\)', texto)
        if anios2:
            fin = anios2.group(2)
            if fin and int(fin) >= anio_actual:
                fin = ""  # abierto: sigue en producción
            m["anios"] = f"{anios2.group(1)}-{fin}" if fin else f"{anios2.group(1)}-"

    # Combustible basado en el nombre del motor
    tl = texto.lower()
    if any(x in tl for x in ["diesel", "diésel", "dci", "hdi", "cdti", "tdi", "d4d", "bluehdi", "d-4d", "tdci"]):
        m["combustible"] = "Diesel"
    elif any(x in tl for x in ["lpg", "glp"]):
        m["combustible"] = "GLP"
    elif any(x in tl for x in ["hybrid", "hibrido", "híbrido", "electric", "eléctrico", "ev", "hev", "phev", "e-tech", "gte"]):
        m["combustible"] = "Hibrido/Electrico"
    else:
        m["combustible"] = "Gasolina"

    # Años (fallback si el bloque anterior no capturó nada)
    if not m["anios"]:
        anios = re.search(r'((?:19|20)\d{2})\s*[-–]\s*((?:19|20)\d{2}|)', texto)
        if anios and int(anios.group(1)) >= 1950:  # sanity check: descartar años absurdos
            fin = anios.group(2)
            if fin and int(fin) >= anio_actual:
                fin = ""
            m["anios"] = f"{anios.group(1)}-{fin}" if fin else f"{anios.group(1)}-"

    if m["codigo"] or m["cv"]:
        return m
    return None


# ─── Excel ────────────────────────────────────────────────────────────────────

def crear_excel(filas: list[dict], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalogo"

    cols = ["Marca", "Modelo", "Variante", "Motor", "Codigo motor",
            "Cilindrada (L)", "CV", "KW", "Combustible", "Anos"]
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    fill_par  = PatternFill("solid", fgColor="EFF6FF")
    fill_impar = PatternFill("solid", fgColor="FFFFFF")

    for ri, fila in enumerate(filas, 2):
        fill = fill_par if ri % 2 == 0 else fill_impar
        vals = [
            fila.get("marca",""), fila.get("modelo",""), fila.get("variante",""),
            fila.get("motor",""), fila.get("codigo",""), fila.get("cilindrada",""),
            fila.get("cv",""), fila.get("kw",""), fila.get("combustible",""),
            fila.get("anios",""),
        ]
        for ci, val in enumerate(vals, 1):
            ws.cell(row=ri, column=ci, value=val).fill = fill

    anchos = [12, 14, 24, 20, 14, 14, 6, 6, 18, 12]
    for ci, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(ci)].width = ancho

    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    print(f"\nExcel guardado: {path} ({len(filas)} filas)")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    todas_filas = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.set_default_timeout(30000)

        for marca in MARCAS:
            print(f"\n{'='*50}")
            print(f"Procesando: {marca.upper()}")
            print(f"{'='*50}")

            modelos = get_modelos(page, marca)
            print(f"  -> {len(modelos)} modelos encontrados")

            for modelo in modelos:
                print(f"\n  Modelo: {modelo['nombre']}")
                time.sleep(DELAY)

                variantes = get_variantes(page, modelo["url"])
                print(f"     -> {len(variantes)} variantes")

                for variante in variantes:
                    print(f"     Variante: {variante['nombre']}", end="", flush=True)
                    time.sleep(DELAY)

                    motores = parse_motores(page, variante["url"])
                    print(f" -> {len(motores)} motores")

                    # Año de la variante: capturado desde la página del modelo (puede estar vacío)
                    # La fuente más fiable es la propia página de la variante (ya cargada en parse_motores)
                    anios_var = get_anios_variante(page)
                    if not anios_var:
                        anios_var = variante.get("anios_variante", "")
                    if motores:
                        for m in motores:
                            # Si el año del motor es inválido (< 1950), usar el año de la variante
                            anios_motor = m.get("anios", "")
                            if anios_motor:
                                inicio = re.match(r'^(\d{4})', anios_motor)
                                if inicio and int(inicio.group(1)) < 1950:
                                    m["anios"] = anios_var
                            todas_filas.append({
                                "marca": marca.capitalize(),
                                "modelo": modelo["nombre"],
                                "variante": variante["nombre"],
                                **m,
                            })
                    else:
                        todas_filas.append({
                            "marca": marca.capitalize(), "modelo": modelo["nombre"],
                            "variante": variante["nombre"], "motor": "", "codigo": "",
                            "cilindrada": "", "cv": "", "kw": "", "combustible": "", "anios": anios_var,
                        })

        browser.close()

    nombre_fichero = f"avtopro_{'_'.join(MARCAS)}.xlsx" if len(MARCAS) > 1 else f"avtopro_{MARCAS[0]}.xlsx"
    crear_excel(todas_filas, nombre_fichero)
    print(f"Total filas: {len(todas_filas)}")


if __name__ == "__main__":
    main()
