"""
Parchea avtopro_nissan.xlsx y avtopro_francesas.xlsx sin re-scrapear.
Correcciones aplicadas:
  1. Código motor: re-extraer desde la descripción del motor
  2. Años: corregir fin de serie cuando == año actual (producción activa)
  3. Modelo: eliminar prefijo "Recambios/Repuestos para [Marca]"
  4. Variante: eliminar prefijo "Recambios/Repuestos [Marca]"
"""
import re
import datetime
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

ANIO_ACTUAL = datetime.date.today().year

FICHEROS = [
    ("avtopro_nissan.xlsx",    "avtopro_nissan_fixed.xlsx"),
    ("avtopro_francesas.xlsx", "avtopro_francesas_fixed.xlsx"),
]


def extraer_codigo(motor_texto: str) -> str:
    """Re-extrae el código de motor desde la descripción completa."""
    pre = re.split(r'\s*\((?:19|20)\d{2}', motor_texto)[0].strip()
    # Dos partes: "F8M 700", "C1E 715", "K9K 608"
    two = re.search(r'\b([A-Z][A-Z0-9]{1,5})\s+(\d{3,4})\s*$', pre)
    if two:
        return two.group(1) + " " + two.group(2)
    # Una parte: "GA14DE", "CD20", "QG16DE", "YD22DDTi", "DV6FC"
    # Permite minúsculas al final pero exige al menos un dígito
    one = re.search(r'\b([A-Z][A-Za-z0-9]{2,8})\s*$', pre)
    if one and re.search(r'\d', one.group(1)):
        return one.group(1)
    return ""


def corregir_anios(anios_str: str) -> str:
    """Corrige el campo de años:
      - Inicio < 1950: dato basura del sitio → vaciar
      - Resto: mantener tal cual (2026 = en producción, es correcto)
    """
    if not anios_str:
        return anios_str
    m = re.match(r'^(\d{4})-(\d{4})?$', str(anios_str))
    if not m:
        return anios_str
    inicio = m.group(1)
    if int(inicio) < 1950:
        return ""  # año de inicio absurdo (ej: 1910) → dato inválido
    return anios_str


def limpiar_modelo(nombre: str) -> str:
    return re.sub(r'^(?:Recambios|Repuestos)\s+para\s+\S+\s+', '', nombre, flags=re.IGNORECASE).strip()


def limpiar_variante(nombre: str) -> str:
    return re.sub(r'^(?:Recambios|Repuestos)\s+\S+\s+', '', nombre, flags=re.IGNORECASE).strip()


def parchear(origen: str, destino: str):
    wb = openpyxl.load_workbook(origen)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data = rows[1:]

    # Índices de columnas (0-based): Marca=0,Modelo=1,Variante=2,Motor=3,Codigo=4,
    # Cilindrada=5,CV=6,KW=7,Combustible=8,Anos=9
    COL_MODELO    = 1
    COL_VARIANTE  = 2
    COL_MOTOR     = 3
    COL_CODIGO    = 4
    COL_ANIOS     = 9

    nuevas_filas = []
    stats = {"codigo_ok": 0, "codigo_vacio": 0, "anios_corr": 0, "modelo_corr": 0, "variante_corr": 0}

    for fila in data:
        fila = list(fila)

        # 1. Código motor
        motor = fila[COL_MOTOR] or ""
        codigo_nuevo = extraer_codigo(motor)
        if codigo_nuevo and not fila[COL_CODIGO]:
            fila[COL_CODIGO] = codigo_nuevo
            stats["codigo_ok"] += 1
        elif not codigo_nuevo:
            stats["codigo_vacio"] += 1

        # 2. Años
        anios_orig = str(fila[COL_ANIOS] or "")
        anios_corr = corregir_anios(anios_orig)
        if anios_corr != anios_orig:
            fila[COL_ANIOS] = anios_corr
            stats["anios_corr"] += 1

        # 3. Nombre modelo
        modelo_orig = str(fila[COL_MODELO] or "")
        modelo_corr = limpiar_modelo(modelo_orig)
        if modelo_corr != modelo_orig:
            fila[COL_MODELO] = modelo_corr
            stats["modelo_corr"] += 1

        # 4. Nombre variante
        var_orig = str(fila[COL_VARIANTE] or "")
        var_corr = limpiar_variante(var_orig)
        if var_corr != var_orig:
            fila[COL_VARIANTE] = var_corr
            stats["variante_corr"] += 1

        nuevas_filas.append(fila)

    # ─── Escribir nuevo Excel ───────────────────────────────────────────────────
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Catalogo"

    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for ci, col in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 20

    fill_par   = PatternFill("solid", fgColor="EFF6FF")
    fill_impar = PatternFill("solid", fgColor="FFFFFF")

    for ri, fila in enumerate(nuevas_filas, 2):
        fill = fill_par if ri % 2 == 0 else fill_impar
        for ci, val in enumerate(fila, 1):
            ws2.cell(row=ri, column=ci, value=val).fill = fill

    anchos = [12, 14, 24, 20, 14, 14, 6, 6, 18, 12]
    for ci, ancho in enumerate(anchos, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = ancho

    ws2.auto_filter.ref = ws2.dimensions
    wb2.save(destino)

    print(f"\n=== {origen} → {destino} ===")
    print(f"  Filas procesadas    : {len(nuevas_filas)}")
    print(f"  Códigos extraídos   : {stats['codigo_ok']}")
    print(f"  Códigos sin extraer : {stats['codigo_vacio']}")
    print(f"  Años corregidos     : {stats['anios_corr']}")
    print(f"  Modelos limpiados   : {stats['modelo_corr']}")
    print(f"  Variantes limpiadas : {stats['variante_corr']}")


if __name__ == "__main__":
    for origen, destino in FICHEROS:
        parchear(origen, destino)
    print("\nListo.")
